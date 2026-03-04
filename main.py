"""
★ 쓰리스타 (Three Star) — main.py v2 ★
PM 전용 파일 — 팀원들은 이 파일을 수정하지 마세요!

[v2 변경사항]
- 직원 포탈 로그인 추가 (routers/auth.py)
- 문의 폼 라우터 추가 (routers/contact.py)
- 포탈 접근 시 로그인 체크 미들웨어 추가
- /product 페이지: products.json 데이터 연결

[v3 변경사항]
- auth.py 세션 기반으로 변경 (쿠키 → 세션)
- 모든 포탈 페이지에 user_name, user_initial 전달
"""

import os
import json
import shutil
from pathlib import Path
from contextlib import asynccontextmanager
from fastapi import FastAPI, Request
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.responses import HTMLResponse, RedirectResponse
from starlette.middleware.base import BaseHTTPMiddleware
import yfinance as yf
import time
from datetime import datetime
import pytz

# ─── 팀원들이 만든 라우터 import ───
from routers import auth
from routers import contact

from routers import recipe_chatbot
from routers import recipe_admin
from slowapi import _rate_limit_exceeded_handler
from slowapi.errors import RateLimitExceeded
from routers.recipe_admin import limiter
from routers import news_summary
from routers import review_keyword
from routers import sns_generator
from routers import market_map
from routers import ideas
from starlette.middleware.sessions import SessionMiddleware
from routers import recipe_analytics
from routers import search

# ─── 서버 시작/종료 이벤트 ───
@asynccontextmanager
async def lifespan(app):
    # ── 시작: ChromaDB 초기화 (삭제 → 재생성) ──
    chroma_path = Path("data/chromadb")
    if chroma_path.exists():
        shutil.rmtree(chroma_path)
        print("[startup] data/chromadb 삭제 완료")

    from services.init_vectordb import main as init_vectordb
    init_vectordb()
    print("[startup] ChromaDB 재생성 완료")

    # recipe_search 싱글턴 초기화
    from services import recipe_search
    recipe_search._collection = None

    yield

# ─── FastAPI 앱 생성 ───
app = FastAPI(
    title="Dami Food Tech Platform",
    description="베트남 K-푸드 브랜드 론칭 · 마케팅 자동화 플랫폼",
    version="0.3.0",
    lifespan=lifespan,
)


# ─── 직원 포탈 인증 미들웨어 ───
class PortalAuthMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        path = request.url.path
        if path.startswith("/portal") and path != "/portal/login":
            if request.session.get("portal_auth") != "authenticated":
                return RedirectResponse("/portal/login", status_code=302)
        return await call_next(request)

app.add_middleware(PortalAuthMiddleware)

app.add_middleware(
    SessionMiddleware,
    secret_key=os.getenv("SESSION_SECRET", "change-me-please"),
    same_site="lax"
)

# slowapi rate limiter
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# 정적 파일 & 템플릿
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")


# ─── 세션에서 유저 정보 가져오는 헬퍼 ───
def get_user(request: Request) -> dict:
    return {
        "user_name":    request.session.get("user_name",    "직원"),
        "user_initial": request.session.get("user_initial", "??"),
        "user_role":    request.session.get("user_role",    "staff"),
    }


# ─── 라우터 연결 ───
app.include_router(auth.router)
app.include_router(contact.router)
app.include_router(recipe_chatbot.router, prefix="/api",          tags=["B-레시피챗봇"])
app.include_router(recipe_admin.router,   prefix="/admin",        tags=["관리자"])
app.include_router(news_summary.router,   prefix="/api",          tags=["C-뉴스요약"])
app.include_router(review_keyword.router, prefix="/api",          tags=["D-리뷰분석"])
app.include_router(sns_generator.router,  prefix="/api",          tags=["E-SNS콘텐츠"])
app.include_router(market_map.router,     prefix="/api",          tags=["F-유통지도"])
app.include_router(ideas.router, prefix="/api/intranet", tags=["인트라넷"])
app.include_router(recipe_analytics.router, prefix="/analytics", tags=["분석"])
app.include_router(search.router,            prefix="/api",          tags=["검색"])

# ═══════════════════════════════════════
# B2C 자사몰
# ═══════════════════════════════════════

@app.get("/", response_class=HTMLResponse, tags=["A-B2C자사몰"])
async def home(request: Request):
    return templates.TemplateResponse("b2c/index.html", {"request": request})


@app.get("/company", response_class=HTMLResponse, tags=["A-B2C자사몰"])
async def company(request: Request):
    return templates.TemplateResponse("b2c/company.html", {"request": request})


@app.get("/product", response_class=HTMLResponse, tags=["A-B2C자사몰"])
async def product(request: Request):
    data = json.loads(Path("data/products.json").read_text(encoding="utf-8"))
    categories = {}
    for p in data["products"]:
        cat = p["category"]
        if cat not in categories:
            categories[cat] = {
                "label":    p["category_label"],
                "label_kr": p["category_label_kr"],
                "products": []
            }
        categories[cat]["products"].append(p)
    return templates.TemplateResponse("b2c/product.html", {
        "request":    request,
        "categories": categories
    })


@app.get("/chatbot", response_class=HTMLResponse, tags=["A-B2C자사몰"])
async def chatbot(request: Request):
    return templates.TemplateResponse("b2c/chatbot.html", {"request": request})


@app.get("/admin-page", response_class=HTMLResponse, tags=["관리자"])
async def admin_page(request: Request):
    return templates.TemplateResponse("admin.html", {"request": request})


# ═══════════════════════════════════════
# B2B 직원 포탈
# ═══════════════════════════════════════

@app.get("/portal", response_class=HTMLResponse, tags=["B2B-직원포탈"])
async def portal_dashboard(request: Request):

    # ── 리뷰 키워드 바: competitor_history.xlsx 에서 최신 데이터 읽기 ──
    dash_keywords = []
    try:
        import openpyxl, re
        xlsx_path = Path("data/competitor_history.xlsx")
        if xlsx_path.exists():
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
            all_rows = list(ws.iter_rows(min_row=2, values_only=True))
            if all_rows:
                latest_text = " ".join([r[2] for r in all_rows if r[2]])
                kw_pattern  = re.findall(r'[🟩🟥🟧⬜]+\s+(.+?)\s+\((\d+)%\)', latest_text)
                seen = {}
                for name, pct in kw_pattern:
                    name = name.strip()
                    if name not in seen and len(name) < 12:
                        seen[name] = int(pct)
                dash_keywords = [{"name": k, "pct": v} for k, v in list(seen.items())[:5]]
    except Exception as e:
        print(f"[대시보드 Excel 읽기 실패] {e}")

    # ── 뉴스 카드: news_cache.json 에서 최신 뉴스 읽기 ──
    dash_news = []
    try:
        news_path = Path("data/news_cache.json")
        if news_path.exists():
            cached   = json.loads(news_path.read_text(encoding="utf-8"))
            dash_news = cached[:5]
    except Exception as e:
        print(f"[대시보드 뉴스 읽기 실패] {e}")

    return templates.TemplateResponse("b2b/dashboard.html", {
        "request":       request,
        "dash_keywords": dash_keywords,
        "dash_news":     dash_news,
        **get_user(request),   # user_name, user_initial, user_role
    })


@app.get("/portal/news", response_class=HTMLResponse, tags=["B2B-직원포탈"])
async def portal_news(request: Request):
    return templates.TemplateResponse("b2b/news.html", {
        "request": request,
        **get_user(request),
    })


@app.get("/portal/review", response_class=HTMLResponse, tags=["B2B-직원포탈"])
async def portal_review(request: Request):
    return templates.TemplateResponse("b2b/review.html", {
        "request": request,
        **get_user(request),
    })


@app.get("/portal/sns", response_class=HTMLResponse, tags=["B2B-직원포탈"])
async def portal_sns(request: Request):
    return templates.TemplateResponse("b2b/sns.html", {
        "request": request,
        **get_user(request),
    })


@app.get("/portal/map")
async def portal_map(request: Request):
    return templates.TemplateResponse("b2b/map.html", {
        "request":        request,
        "google_maps_key": os.getenv("GOOGLE_MAPS_API_KEY"),
        **get_user(request),
    })


@app.get("/portal/ideas", response_class=HTMLResponse, tags=["B2B-직원포탈"])
async def portal_ideas(request: Request):
    return templates.TemplateResponse("b2b/ideas.html", {
        "request": request,
        **get_user(request),
    })


@app.get("/portal/customers", response_class=HTMLResponse, tags=["B2B-직원포탈"])
async def portal_customers(request: Request):
    return templates.TemplateResponse("b2b/customers.html", {
        "request": request,
        "active_menu": "customers",
        **get_user(request),
    })


@app.get("/portal/api/analytics", tags=["B2B-직원포탈"])
async def portal_analytics(days: int = 7):
    """포탈용 분석 API — 세션 쿠키 인증 (미들웨어가 /portal/* 자동 보호)"""
    from services.analytics import (
        get_daily_stats, get_popular_recipes, get_cart_ranking,
        get_taste_trends, get_language_stats,
    )
    recipes_dir = Path("data/recipes")
    total_recipes = len(list(recipes_dir.glob("recipe_*.md"))) if recipes_dir.exists() else 0

    return {
        "daily":         get_daily_stats(days),
        "popular":       get_popular_recipes(days),
        "cart":          get_cart_ranking(days),
        "trends":        get_taste_trends(days),
        "languages":     get_language_stats(days),
        "total_recipes": total_recipes,
    }


@app.get("/portal/report", response_class=HTMLResponse, tags=["B2B-직원포탈"])
async def portal_report(request: Request):

    # ── 뉴스 데이터 ──
    news_items = []
    try:
        news_path = Path("data/news_cache.json")
        if news_path.exists():
            cached     = json.loads(news_path.read_text(encoding="utf-8"))
            news_items = cached[:8]
    except Exception as e:
        print(f"[리포트 뉴스 읽기 실패] {e}")

    # ── 키워드 데이터 ──
    report_keywords = []
    try:
        import openpyxl, re
        xlsx_path = Path("data/competitor_history.xlsx")
        if xlsx_path.exists():
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
            all_rows = list(ws.iter_rows(min_row=2, values_only=True))
            if all_rows:
                latest_text = " ".join([r[2] for r in all_rows if r[2]])
                kw_pattern  = re.findall(r'[🟩🟥🟧⬜]+\s+(.+?)\s+\((\d+)%\)', latest_text)
                seen = {}
                for name, pct in kw_pattern:
                    name = name.strip()
                    if name not in seen and len(name) < 12:
                        seen[name] = int(pct)
                report_keywords = [{"name": k, "pct": v} for k, v in list(seen.items())[:8]]
    except Exception as e:
        print(f"[리포트 Excel 읽기 실패] {e}")

    from datetime import datetime, timedelta
    today      = datetime.now()
    week_start = (today - timedelta(days=today.weekday())).strftime("%Y.%m.%d")
    week_end   = today.strftime("%Y.%m.%d")

    return templates.TemplateResponse("b2b/report.html", {
        "request":         request,
        "news_items":      news_items,
        "report_keywords": report_keywords,
        "week_start":      week_start,
        "week_end":        week_end,
        "generated_at":    today.strftime("%Y.%m.%d %H:%M"),
        **get_user(request),
    })


# ── 환율 API (yfinance + 1시간 캐싱 적용) ──────────────────────────────────────
_exchange_cache = {}
_CACHE_TTL = 3600  # 1시간(3600초) 캐싱

@app.get("/api/exchange-rate", tags=["유틸"])
async def exchange_rate():
    current_time = time.time()
    
    # 1. 캐시 확인 (1시간 이내면 바로 반환)
    if "rates" in _exchange_cache and (current_time - _exchange_cache["last_updated"] < _CACHE_TTL):
        return _exchange_cache["rates"]

    try:
        # 2. yfinance로 실시간 데이터 가져오기
        krw_ticker = yf.Ticker("KRW=X")
        vnd_ticker = yf.Ticker("VND=X")
        
        krw_price = krw_ticker.history(period="1d")['Close'].iloc[-1]
        vnd_price = vnd_ticker.history(period="1d")['Close'].iloc[-1]

        # 환율 계산
        vnd_to_krw = krw_price / vnd_price
        vnd100_krw = vnd_to_krw * 100

        # 서울 기준 시간 생성
        seoul_tz = pytz.timezone('Asia/Seoul')
        updated_time = datetime.now(seoul_tz).strftime("%Y.%m.%d %H:00")

        # 프론트엔드와 형태 맞추기
        rates_data = {
            "success": True,
            "usd_krw": round(krw_price, 2),
            "usd_vnd": round(vnd_price, 2),
            "vnd_to_krw": round(vnd_to_krw, 6),
            "vnd100_krw": round(vnd100_krw, 4),
            "updated": updated_time,
        }

        # 3. 캐시 저장
        _exchange_cache["rates"] = rates_data
        _exchange_cache["last_updated"] = current_time

        return rates_data

    except Exception as e:
        print(f"[환율 API 오류] {e}")
        return {
            "success": False, 
            "error": str(e),
            "usd_krw": 1330.00,
            "vnd100_krw": 5.40,
            "updated": "기본값 (오류)"
        }

# ── 서버 실행 (Render 배포용 포트 설정) ────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn
    # Render는 0.0.0.0 호스트와 환경변수 PORT를 사용해야 정상 작동합니다.
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run("main:app", host="127.0.0.1", port=port)